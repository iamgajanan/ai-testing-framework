from __future__ import annotations

import csv
import json
import mimetypes
import re
from pathlib import Path
from typing import Any


def _resolve_path(path: str, base_dir: str | Path | None = None) -> Path:
    candidate = Path(path).expanduser()
    if not candidate.is_absolute() and base_dir:
        candidate = Path(base_dir) / candidate
    return candidate.resolve()


def _load_content(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _validate_json(path: Path, expected: Any = None, json_path: str = "") -> tuple[bool, str]:
    try:
        data = json.loads(_load_content(path))
    except (OSError, ValueError) as exc:
        return False, f"Invalid JSON: {exc}"
    if json_path:
        current = data
        try:
            for part in json_path.split("."):
                if isinstance(current, dict):
                    current = current[part]
                elif isinstance(current, list) and part.isdigit():
                    current = current[int(part)]
                else:
                    raise KeyError(part)
        except (KeyError, IndexError, TypeError):
            return False, f"JSON path not found: {json_path}"
        if expected is not None and current != expected:
            return False, f"JSON {json_path!r} expected {expected!r}, got {current!r}"
    return True, "JSON file is valid"


def _validate_csv(path: Path, expected_columns: list[str] | None = None) -> tuple[bool, str]:
    try:
        with path.open(newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            columns = reader.fieldnames or []
            missing = [c for c in (expected_columns or []) if c not in columns]
            if missing:
                return False, f"CSV missing expected columns: {', '.join(missing)}"
            list(reader)
    except (OSError, csv.Error) as exc:
        return False, f"Invalid CSV: {exc}"
    return True, "CSV file is valid"


def _validate_xlsx(path: Path, expected_columns: list[str] | None = None) -> tuple[bool, str]:
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return False, f"Invalid XLSX: {exc}"
    try:
        sheet = workbook.active
        first = next(sheet.iter_rows(values_only=True), ())
        columns = [str(v).strip() for v in first if v is not None]
        missing = [c for c in (expected_columns or []) if c not in columns]
        if missing:
            return False, f"XLSX missing expected columns: {', '.join(missing)}"
    finally:
        workbook.close()
    return True, "XLSX file is valid"


def validate_file(
    path: str,
    *,
    base_dir: str | Path | None = None,
    expected_filename: str | None = None,
    expected_extension: str | None = None,
    expected_mime: str | None = None,
    min_size: int | None = None,
    max_size: int | None = None,
    text_contains: str | None = None,
    pattern: str | None = None,
    file_type: str | None = None,
    json_path: str | None = None,
    expected: Any = None,
    expected_columns: list[str] | None = None,
) -> tuple[bool, str, dict[str, Any]]:
    target = _resolve_path(path, base_dir)
    meta: dict[str, Any] = {"path": str(target)}
    if not target.exists():
        return False, f"File does not exist: {target}", meta
    if not target.is_file():
        return False, f"Path is not a file: {target}", meta

    stat = target.stat()
    mime = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
    meta.update({"filename": target.name, "extension": target.suffix.lower(), "mime": mime, "size": stat.st_size})

    checks: list[str] = []
    if expected_filename and target.name != expected_filename:
        return False, f"Filename expected {expected_filename!r}, got {target.name!r}", meta
    if expected_extension and target.suffix.lower() != expected_extension.lower():
        return False, f"Extension expected {expected_extension!r}, got {target.suffix!r}", meta
    if expected_mime and not (mime == expected_mime or mime.startswith(expected_mime.rstrip("/*") + "/")):
        return False, f"MIME expected {expected_mime!r}, got {mime!r}", meta
    if min_size is not None and stat.st_size < min_size:
        return False, f"File size {stat.st_size} is below minimum {min_size}", meta
    if max_size is not None and stat.st_size > max_size:
        return False, f"File size {stat.st_size} exceeds maximum {max_size}", meta

    kind = (file_type or target.suffix.lstrip(".")).lower()
    if kind in {"json"}:
        ok, reason = _validate_json(target, expected, json_path or "")
        if not ok:
            return False, reason, meta
        checks.append(reason)
    elif kind in {"csv"}:
        ok, reason = _validate_csv(target, expected_columns)
        if not ok:
            return False, reason, meta
        checks.append(reason)
    elif kind in {"xlsx", "xlsm", "excel"}:
        ok, reason = _validate_xlsx(target, expected_columns)
        if not ok:
            return False, reason, meta
        checks.append(reason)
    elif kind in {"pdf"}:
        try:
            from pypdf import PdfReader
            PdfReader(str(target))
            checks.append("PDF structure is valid")
        except ImportError:
            # Magic-header validation keeps PDF checking available without adding a hard dependency.
            if target.read_bytes()[:5] != b"%PDF-":
                return False, "File is not a valid PDF (missing %PDF header)", meta
            checks.append("PDF header is valid")
        except Exception as exc:
            return False, f"Invalid PDF: {exc}", meta

    if text_contains or pattern:
        try:
            text = _load_content(target)
        except OSError as exc:
            return False, f"Could not read file as text: {exc}", meta
        if text_contains and text_contains.lower() not in text.lower():
            return False, f"File content does not contain {text_contains!r}", meta
        if pattern:
            try:
                if not re.search(pattern, text, re.MULTILINE):
                    return False, f"File content does not match pattern {pattern!r}", meta
            except re.error as exc:
                return False, f"Invalid file regex: {exc}", meta
        checks.append("content assertions passed")

    return True, "; ".join(checks) if checks else "File validation passed", meta
