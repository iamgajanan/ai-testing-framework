from __future__ import annotations

from pathlib import Path
from typing import Any

from ..core.models import Step, TestCase, TestSuite, Validation
from .csv_parser import CSVParser


class XLSXParser:
    """Parse the first worksheet of an Excel test suite."""

    def parse(self, path: str | Path) -> TestSuite:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("XLSX support requires openpyxl. Install it with: pip install openpyxl") from exc
        workbook = load_workbook(Path(path), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value).strip() if value is not None else "" for value in next(rows, ())]
            tests: dict[str, TestCase] = {}
            parser = CSVParser()
            for values in rows:
                row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]}
                parser._add_row(tests, row)
            if not tests:
                raise ValueError("XLSX suite must contain at least one row with TestID.")
            return TestSuite("XLSX Test Suite", list(tests.values()))
        finally:
            workbook.close()
